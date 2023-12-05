import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import calendar

# Initialize the workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Define the year and the days of the week
year = 2024
days_of_week = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Define fill styles for alternating weeks
fill_colors = ["FF9999", "9999FF"]
fills = [
    PatternFill(start_color=color, end_color=color, fill_type="solid")
    for color in fill_colors
]


# Function to add a month to the sheet
def add_month_to_sheet(sheet, year, month, start_row, start_col):
    # Merge cells for the month header
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 6,
    )
    month_cell = sheet.cell(row=start_row, column=start_col)
    month_cell.value = calendar.month_name[month]
    month_cell.alignment = Alignment(horizontal="center")

    # Fill in the days of the week headers
    for i, day in enumerate(days_of_week, start=start_col):
        cell = sheet.cell(row=start_row + 1, column=i)
        cell.value = day
        cell.alignment = Alignment(horizontal="center")

    # Add the days for each week of the month
    week_fill = 0  # Start with the first color
    for week_idx, week in enumerate(calendar.monthcalendar(year, month)):
        for i, day in enumerate(week, start=start_col):
            cell = sheet.cell(row=start_row + 2 + week_idx, column=i)
            cell.alignment = Alignment(horizontal="center")
            if day != 0:  # Avoid inserting 0 for days that are not part of the month
                cell.value = day
                cell.fill = fills[week_fill]
        week_fill = (week_fill + 1) % 2  # Alternate the fill color


# Add the months in sets of three
row_start = 1
for quarter in range(0, 12, 3):
    max_weeks = 0
    for month_offset in range(3):
        month = quarter + month_offset + 1
        start_col = month_offset * 9 + 1  # 7 days + 2 spacing columns
        add_month_to_sheet(sheet, year, month, row_start, start_col)
        weeks_in_month = len(calendar.monthcalendar(year, month))
        if weeks_in_month > max_weeks:
            max_weeks = weeks_in_month
    row_start += max_weeks + 3  # Add extra space after each set of 3 months

# Auto-adjust column widths
for col in sheet.columns:
    max_length = max(
        len(str(cell.value)) if cell.value is not None else 0 for cell in col
    )
    adjusted_width = max_length + 2
    sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Save the workbook to a file
file_path = "2024_Calendar.xlsx"
wb.save(file_path)

file_path
